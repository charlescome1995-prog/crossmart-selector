# -*- coding: utf-8 -*-
"""
选品模块 - 完整版（基于16个核心指标 + 5W1H场景还原）

核心标准（来自2025亚马逊爆款选品思维）:
1. 价格$10-50（冲动购买甜蜜点）
2. 月销300-900（10×10×1法则）
3. 竞品评论<10,000（避免饱和）
4. 毛利率>30%红线
5. 无大品牌垄断
6. 非季节性
7. 市场深度（前3<60%）

4种策略:
  策略1_蓝海策略: 小众市场、低竞争、高利润
  策略2_红海策略: 成熟大市场、高销量
  策略3_差异化策略: 细分市场、有改进空间
  策略4_跟随策略: 已验证市场、稳定收益

用法:
  cd 1data_cl_selection
  python product_selector.py
"""
import sys
import os

_this_dir = os.path.dirname(os.path.abspath(__file__))
_project_dir = os.path.dirname(_this_dir)
if _project_dir not in sys.path:
    sys.path.insert(0, _project_dir)
if _this_dir not in sys.path:
    sys.path.insert(0, _this_dir)

sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import numpy as np
from datetime import datetime
import json

from config import RAW_DATA_DIR, OUTPUT_DIR, STRATEGY_CONFIG, SELECTION_RULES, PROFIT_MODEL, BASE_DIR


# ============================================================
# 16个核心指标检测
# ============================================================

def check_16_indicators(row):
    """
    检测产品是否符合16个核心选品指标
    返回: dict，包含每个指标的通过情况
    """
    indicators = {}
    
    # 1. 价格介于$10-50
    price = row.get('价格（美元）', 0)
    indicators['price_10_50'] = 10 <= price <= 50
    indicators['price_optimal'] = 20 <= price <= 50
    
    # 2. 重量（数据中无此字段，标记为N/A）
    indicators['weight_light'] = True  # 假设通过
    
    # 3. 大分类排名<5000
    rank = row.get('rank_latest', 0)
    indicators['rank_top5000'] = 0 < rank <= 5000
    indicators['rank_top10000'] = 0 < rank <= 10000
    
    # 4. 无大品牌垄断
    keyword = str(row.get('关键词(绿色建议进入，黄色找切入点，粉色观察)', '')).lower()
    brand_keywords = SELECTION_RULES.get('brand_monopoly_keywords', [])
    indicators['no_brand_monopoly'] = not any(b in keyword for b in brand_keywords)
    
    # 5. 坚固耐用（数据中无此字段，假设通过）
    indicators['durable'] = True
    
    # 6. 首页2-3个Review<50（通过竞品分析判断）
    review_count = row.get('评分数', 0)
    indicators['low_review_opportunity'] = review_count < 5000
    
    # 7. 成本<售价25%（毛利率>30%）
    gross_margin = calculate_profit_margin(price)
    indicators['margin_30'] = gross_margin >= 30
    
    # 8. Listing有优化空间
    rating = row.get('评分值', 0)
    indicators['optimization_space'] = 0 < rating < 4.6
    
    # 9. 前三关键字月搜>100,000
    monthly_search = row.get('20260113月搜', 0)
    indicators['high_search_volume'] = monthly_search >= 10000  # 放宽到10K
    
    # 10. 中国可找到供货商（假设通过）
    indicators['china_supply'] = True
    
    # 11. 非季节性
    indicators['non_seasonal'] = not is_seasonal(keyword)
    
    # 12. eBay也有售（假设通过）
    indicators['multi_platform'] = True
    
    # 13. 可拓展品牌
    indicators['brand_extensible'] = True
    
    # 14. 可升级改良
    indicators['upgradeable'] = indicators['optimization_space']
    
    # 15. 需定期购买
    indicators['repeat_purchase'] = True
    
    # 16. 多样化关键词
    indicators['keyword_diversity'] = len(keyword.split()) >= 2
    
    # 计算通过率
    passed = sum([
        indicators['price_10_50'],
        indicators['rank_top10000'],
        indicators['no_brand_monopoly'],
        indicators['low_review_opportunity'],
        indicators['margin_30'],
        indicators['optimization_space'],
        indicators['high_search_volume'],
        indicators['non_seasonal'],
        indicators['keyword_diversity'],
    ])
    indicators['pass_rate'] = passed / 9 * 100
    
    return indicators


def calculate_profit_margin(price, weight_kg=0.3, is_small_standard=True):
    """
    计算预期毛利率（基于亚马逊FBA官方费用表2025-2026）
    
    参数：
    - price: 售价（美元）
    - weight_kg: 产品重量（公斤），默认300g
    - is_small_standard: 是否Small Standard尺寸（最优）
    
    返回：毛利率百分比
    """
    if price <= 0:
        return 0
    
    # 1. 产品成本（供应链优化后可降到15%）
    product_cost = price * PROFIT_MODEL['product_cost_ratio']
    
    # 2. 亚马逊佣金（大多数类目15%）
    commission = price * PROFIT_MODEL['amazon_commission_ratio']
    
    # 3. FBA费用（基于Size Tier）
    if is_small_standard:
        fba_fee = PROFIT_MODEL['fba_small_standard']
    else:
        # 根据重量判断Large Standard子类
        weight_lb = weight_kg * 2.20462
        if weight_lb <= 1:
            fba_fee = PROFIT_MODEL['fba_large_standard_light']
        elif weight_lb <= 3:
            fba_fee = PROFIT_MODEL['fba_large_standard_medium']
        else:
            fba_fee = PROFIT_MODEL['fba_large_standard_heavy']
    
    # 4. 头程运费（空运，重量×单价）
    first_shipping = weight_kg * PROFIT_MODEL['first_mile_air_per_kg']
    
    # 5. 广告成本（新品期12%）
    ad_cost = price * PROFIT_MODEL['ad_cost_ratio_new']
    
    # 6. 其他成本（退货损耗、仓储等）
    return_cost = price * PROFIT_MODEL['return_rate'] * 0.5  # 退货只损失一半售价（可转售）
    storage_cost = PROFIT_MODEL['storage_fee_monthly']
    other_cost = price * PROFIT_MODEL['other_cost_ratio']
    
    # 总成本
    total_cost = (
        product_cost + 
        commission + 
        fba_fee + 
        first_shipping + 
        ad_cost + 
        return_cost + 
        storage_cost + 
        other_cost
    )
    
    # 毛利
    gross_profit = price - total_cost
    
    # 毛利率
    margin = (gross_profit / price) * 100 if price > 0 else 0
    
    return round(margin, 1)


def calculate_profit_per_unit(price, margin):
    """计算单件利润（美元）"""
    return round(price * margin / 100, 2)


def is_seasonal(keyword):
    """检测是否季节性产品"""
    keyword_lower = keyword.lower()
    seasonal_keywords = SELECTION_RULES.get('seasonal_blacklist', [])
    return any(kw in keyword_lower for kw in seasonal_keywords)


def is_trending(keyword):
    """检测是否热点事件驱动"""
    keyword_lower = keyword.lower()
    trending_keywords = SELECTION_RULES.get('trending_keywords', [])
    return any(kw in keyword_lower for kw in trending_keywords)


# ============================================================
# 类别优先级 + 食品过滤 + 补充剂标注
# ============================================================

# 直接过滤：食品类
FOOD_KEYWORDS = [
    'food', 'snack', 'candy', 'chocolate', 'cookie', 'cereal', 'milk', 'juice',
    'coffee bean', 'tea bag', 'protein bar', 'energy drink', 'sauce', 'seasoning',
    'spice', 'flour', 'sugar', 'oil', 'vinegar', 'jam', 'honey', 'nut butter',
    'chips', 'popcorn', 'cracker', 'gummy', 'drink mix'
]

# 谨慎考虑：补充剂类（需核查审核门槛）
SUPPLEMENT_KEYWORDS = [
    'supplement', 'vitamin', 'collagen', 'probiotic', 'omega', 'protein powder',
    'whey', 'creatine', 'bcaa', 'pre workout', 'multivitamin', 'fish oil',
    'magnesium', 'zinc', 'iron supplement', 'calcium supplement', 'melatonin',
    'elderberry', 'turmeric capsule', 'ashwagandha', 'maca root'
]

# 高优先：美妆/个护类
BEAUTY_KEYWORDS = [
    'nail', 'skincare', 'skin care', 'serum', 'moisturizer', 'sunscreen', 'spf',
    'face mask', 'eye cream', 'toner', 'cleanser', 'exfoliant', 'retinol',
    'hair', 'hairbrush', 'comb', 'scalp', 'shampoo', 'conditioner', 'hair mask',
    'dry shampoo', 'hair oil', 'hair growth', 'eyelash', 'eyebrow', 'lip balm',
    'deodorant', 'body wash', 'lotion', 'body butter', 'self tan', 'spray tan',
    'makeup', 'foundation', 'concealer', 'blush', 'highlighter', 'setting spray',
    'brush set', 'beauty blender', 'jade roller', 'gua sha', 'facial roller',
    'electric nail', 'nail clipper', 'nail file', 'cuticle', 'pedicure'
]

# 次优先：垄断性不强的电子产品
ELECTRONICS_KEYWORDS = [
    'led light', 'night light', 'desk lamp', 'ring light', 'grow light',
    'bluetooth speaker', 'wireless charger', 'usb hub', 'cable organizer',
    'phone stand', 'tablet stand', 'monitor stand', 'webcam', 'microphone',
    'keyboard', 'mouse pad', 'laptop stand', 'power bank', 'solar charger',
    'smart plug', 'extension cord', 'surge protector', 'cable management',
    'electric', 'rechargeable', 'cordless', 'portable'
]


def get_category_priority(keyword):
    """
    判断产品类别优先级
    返回: (优先级标签, 备注)
    """
    keyword_lower = keyword.lower()

    # 食品类 → 直接过滤
    if any(w in keyword_lower for w in FOOD_KEYWORDS):
        return "❌过滤", "食品类不做"

    # 补充剂类 → 谨慎
    if any(w in keyword_lower for w in SUPPLEMENT_KEYWORDS):
        return "⚠️谨慎", "补充剂/需核查审核门槛"

    # 美妆/个护 → 高优先
    if any(w in keyword_lower for w in BEAUTY_KEYWORDS):
        return "🌟高优先", "美妆/个护"

    # 电子产品（垄断性不强）→ 次优先
    if any(w in keyword_lower for w in ELECTRONICS_KEYWORDS):
        return "⭐次优先", "电子产品"

    return "普通", ""


# ============================================================
# 5W1H场景还原分析（基于"方法A"）
# ============================================================

def analyze_5w1h_scene(keyword, monthly_search, price, review_count, rating):
    """
    5W1H场景还原分析
    Who-谁: 用户是谁
    What-什么: 用户需要什么
    When-何时: 使用场景
    Where-何地: 使用地点
    Why-为何: 购买动机
    How-如何: 如何解决问题
    """
    keyword_lower = keyword.lower()
    
    scene = {
        "who": "未知用户群体",
        "what": "未知需求",
        "when": "日常使用",
        "where": "家庭/室内",
        "why": "解决日常问题",
        "how": "待分析",
        "pain_points": [],
        "opportunities": [],
        "scene_description": ""
    }
    
    # 基于关键词类型推断场景
    if any(w in keyword_lower for w in ['baby', 'infant', 'toddler', 'kids', 'children']):
        scene["who"] = "新生儿/幼儿家长"
        scene["where"] = "婴儿房/客厅"
        scene["pain_points"] = ["照顾婴儿耗时费力", "寻找安全产品困难", "担心产品质量"]
        
    elif any(w in keyword_lower for w in ['kitchen', 'cooking', 'baking', 'food']):
        scene["who"] = "家庭厨师/烘焙爱好者"
        scene["where"] = "厨房"
        scene["pain_points"] = ["厨房收纳困难", "烹饪工具不好用", "清洁麻烦"]
        
    elif any(w in keyword_lower for w in ['office', 'desk', 'computer', 'keyboard']):
        scene["who"] = "办公室职员/远程工作者"
        scene["where"] = "办公室/家庭书房"
        scene["pain_points"] = ["久坐不适", "桌面杂乱", "效率低下"]
        
    elif any(w in keyword_lower for w in ['bathroom', 'shower', 'towel', 'toilet']):
        scene["who"] = "注重生活品质的家庭"
        scene["where"] = "浴室/卫生间"
        scene["pain_points"] = ["浴室潮湿", "收纳空间不足", "清洁困难"]

    elif any(w in keyword_lower for w in ['nail', 'skincare', 'skin care', 'serum', 'moisturizer',
                                           'sunscreen', 'face mask', 'eye cream', 'retinol',
                                           'cleanser', 'toner', 'exfoliant', 'lip balm',
                                           'makeup', 'foundation', 'concealer', 'blush',
                                           'brush set', 'beauty blender', 'jade roller', 'gua sha',
                                           'electric nail', 'nail clipper', 'cuticle', 'pedicure',
                                           'eyelash', 'eyebrow', 'setting spray']):
        scene["who"] = "注重外表的女性/美妆爱好者"
        scene["where"] = "家中/浴室/梳妆台"
        scene["pain_points"] = ["护肤步骤繁琐", "产品效果不持久", "工具不好用/难清洁", "专业美容太贵"]
        scene["opportunities"] = ["居家美容需求增长", "工具类竞品评分普遍偏低", "差异化空间大"]

    elif any(w in keyword_lower for w in ['hair', 'hairbrush', 'comb', 'scalp', 'shampoo',
                                           'conditioner', 'hair mask', 'dry shampoo', 'hair oil',
                                           'hair growth', 'hair loss']):
        scene["who"] = "注重发质的女性/脱发困扰人群"
        scene["where"] = "浴室/梳妆台"
        scene["pain_points"] = ["发梳难清洁", "掉发严重", "头皮问题", "发质受损"]
        scene["opportunities"] = ["自清洁发梳是高频痛点", "头皮护理市场快速增长"]

    elif any(w in keyword_lower for w in ['deodorant', 'body wash', 'lotion', 'body butter',
                                           'self tan', 'spray tan', 'body scrub', 'body oil']):
        scene["who"] = "注重个人卫生/外表的消费者"
        scene["where"] = "浴室/家中"
        scene["pain_points"] = ["气味管理困难", "皮肤干燥", "美黑效果不自然"]
        scene["opportunities"] = ["多功能合一产品受欢迎", "天然成分需求增长"]
        
    elif any(w in keyword_lower for w in ['fitness', 'exercise', 'gym', 'workout']):
        scene["who"] = "健身爱好者"
        scene["where"] = "健身房/家庭"
        scene["pain_points"] = ["缺乏动力", "设备占用空间大", "噪音问题"]
        
    elif any(w in keyword_lower for w in ['pet', 'dog', 'cat', 'animal']):
        scene["who"] = "宠物主人"
        scene["where"] = "家中/户外遛狗"
        scene["pain_points"] = ["宠物管理困难", "清洁麻烦", "宠物安全担忧"]
        
    elif any(w in keyword_lower for w in ['travel', 'camping', 'outdoor']):
        scene["who"] = "旅行/露营爱好者"
        scene["where"] = "户外/旅途"
        scene["pain_points"] = ["携带不便", "空间有限", "需要多功能产品"]
        
    else:
        scene["who"] = "普通消费者"
        scene["where"] = "家庭/室内"
        scene["pain_points"] = ["寻找性价比产品", "担心质量问题", "难以选择"]
    
    # 分析市场机会
    if rating < 4.3:
        scene["opportunities"].append("当前产品评分偏低，存在改进空间")
    if review_count < 3000:
        scene["opportunities"].append("竞品评论数较少，新进入者有机会")
    if price < 30:
        scene["opportunities"].append("价格适中，适合冲动购买")
    if monthly_search > 20000:
        scene["opportunities"].append("搜索量大，市场需求旺盛")
    
    # 生成场景描述
    scene["scene_description"] = f"{scene['who']}在{scene['where']}使用产品来{scene['why']}，主要痛点包括：{'、'.join(scene['pain_points'][:2])}"
    
    return scene


# ============================================================
# 大模型API调用（DeepSeek/Kimi/Gemini）
# ============================================================

def call_llm_for_analysis(keyword, row_data=None):
    """
    调用大模型分析季节性、热点事件、品牌竞争
    返回详细的分析结果
    
    分析内容：
    1. 季节性：3个月后是否旺季
    2. 热点事件：3个月后热度是否仍高
    3. 品牌竞争：是否是新卖家可以利用的窗口期
    """
    from config import AI_CONFIG
    
    # 获取API配置
    provider = AI_CONFIG.get('default_provider', 'deepseek')
    api_config = AI_CONFIG.get(provider, {})
    
    if not api_config.get('enabled', False):
        # 如果API未启用，使用关键词匹配
        return simple_keyword_analysis(keyword)
    
    api_key = api_config.get('api_key', '')
    base_url = api_config.get('base_url', '')
    model = api_config.get('model', 'deepseek-chat')
    
    # 构建提示词
    prompt = f"""分析以下亚马逊产品关键词的选品可行性：

关键词：{keyword}

请分析以下4个维度的风险和机会（每个维度100字以内）：

1. **季节性风险**：
   - 这个关键词是否季节性产品？
   - 如果是，3个月后（6月底）是否进入旺季？
   - 新卖家现在进入是否有时间窗口？

2. **热点事件**：
   - 这个关键词是否受热点事件驱动（如TikTok爆款、节日、新闻等）？
   - 如果是热点，3个月后热度预计会如何变化？
   - 现在进入是否还有机会？

3. **品牌竞争**：
   - 该品类是否存在强势品牌垄断？
   - 如果有强势品牌，新进入者是否有差异化机会？
   - 该关键词是否处于品牌格局变化期（窗口期）？

4. **差异化切入点**（核心）：
   - 从【人群法】角度：这个产品主要服务哪类人群？有没有被忽视的细分人群？
   - 从【场景法】角度：有哪些高频使用场景？竞品没覆盖到的场景是什么？
   - 从【竞品对比】角度：现有竞品普遍的短板/差评痛点可能是什么？新卖家可以从哪个具体功能或卖点切入？
   - 给出1个最可执行的差异化切入建议（如"主打XX人群+XX场景，强化XX功能"）。

请用以下JSON格式返回（只返回JSON，不要其他内容）：
{{
    "is_seasonal": true/false,
    "seasonal_detail": "季节性分析原因（50字以内）",
    "seasonal_window": "3个月后是否旺季，窗口期分析（30字以内）",
    "is_trending": true/false,
    "trending_detail": "热点事件分析原因（50字以内）",
    "trending_future": "3个月后热度预测（30字以内）",
    "has_brand_competition": true/false,
    "brand_detail": "品牌竞争分析原因（50字以内）",
    "brand_window": "是否窗口期，分析原因（30字以内）",
    "target_audience": "核心人群+被忽视的细分人群（40字以内）",
    "key_scenarios": "高频场景+竞品未覆盖场景（40字以内）",
    "competitor_gap": "竞品普遍短板/差评痛点（40字以内）",
    "entry_suggestion": "1个最可执行的差异化切入建议（50字以内）",
    "overall_reason": "综合分析理由（100字以内）"
}}"""
    
    try:
        import requests
        import time as _time
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        data = {
            "model": model,
            "messages": [
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.3,
            "max_tokens": 2000
        }
        
        # 火山方舟偏发限流(429)/长 prompt 偶发空返回，最多重试 3 次
        content = ''
        response = None
        for _attempt in range(3):
            response = requests.post(
                f"{base_url}/chat/completions",
                headers=headers,
                json=data,
                timeout=40
            )
            if response.status_code == 429:
                _time.sleep(3 + _attempt * 3)  # 退避重试
                continue
            if response.status_code == 200:
                content = (response.json()['choices'][0]['message']['content'] or '').strip()
                if content:
                    break
                _time.sleep(2)  # 空返回，稍等重试
            else:
                break
        
        if response is not None and response.status_code == 200 and content:
            import json
            import re
            
            # 提取JSON
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                analysis = json.loads(json_match.group())
                
                # 生成分析理由
                reasons = []
                if analysis.get('is_seasonal'):
                    reasons.append(f"🌡️季节:{analysis.get('seasonal_window', '')}")
                if analysis.get('is_trending'):
                    reasons.append(f"🔥热点:{analysis.get('trending_future', '')}")
                if analysis.get('has_brand_competition'):
                    reasons.append(f"🏢品牌:{analysis.get('brand_window', '')}")
                if analysis.get('entry_suggestion'):
                    reasons.append(f"🎯切入:{analysis.get('entry_suggestion', '')}")
                if not reasons:
                    reasons.append("✅常规产品")
                
                analysis['reason'] = " | ".join(reasons)
                return analysis
    except Exception as e:
        print(f"    LLM分析失败: {str(e)[:50]}, 使用关键词匹配...")
    
    # 备用：使用关键词匹配
    return simple_keyword_analysis(keyword)


def simple_keyword_analysis(keyword):
    """简单的关键词匹配分析（备用）"""
    keyword_lower = keyword.lower()
    
    # 季节性检测
    seasonal_keywords = SELECTION_RULES.get('seasonal_blacklist', [])
    is_seasonal = any(kw in keyword_lower for kw in seasonal_keywords)
    
    # 热点检测
    trending_keywords = SELECTION_RULES.get('trending_keywords', [])
    is_trending = any(kw in keyword_lower for kw in trending_keywords)
    
    # 品牌垄断检测
    brand_keywords = SELECTION_RULES.get('brand_monopoly_keywords', [])
    has_brand = any(kw in keyword_lower for kw in brand_keywords)
    
    # 生成分析理由
    reasons = []
    if is_seasonal:
        reasons.append("⚠️季节性产品")
    if is_trending:
        reasons.append("🔥热点事件驱动")
    if has_brand:
        reasons.append("⚠️可能存在品牌竞争")
    
    if not reasons:
        reasons.append("✅常规产品")
    
    return {
        "is_seasonal": is_seasonal,
        "seasonal_detail": "通过关键词判断为季节性产品",
        "seasonal_window": "需进一步分析旺季时间",
        "is_trending": is_trending,
        "trending_detail": "通过关键词判断为热点驱动",
        "trending_future": "需进一步分析热度趋势",
        "has_brand_competition": has_brand,
        "brand_detail": "通过关键词判断可能存在品牌竞争",
        "brand_window": "需进一步分析品牌格局",
        "overall_reason": " | ".join(reasons),
        "reason": " | ".join(reasons)
    }


# ============================================================
# 筛选函数
# ============================================================

def check_strategy_match(row, strategy_key):
    """检查产品是否符合策略条件"""
    cfg = STRATEGY_CONFIG[strategy_key]
    
    # 排名检查
    rank_earliest = row.get('rank_earliest', 0)
    rank_latest = row.get('rank_latest', 0)
    rank_change_rate = row.get('rank_change_rate', 0)
    
    if 'rank_earliest' in cfg:
        if 'min' in cfg['rank_earliest'] and rank_earliest < cfg['rank_earliest']['min']:
            return False
        if 'max' in cfg['rank_earliest'] and rank_earliest > cfg['rank_earliest']['max']:
            return False
    
    if 'rank_latest' in cfg:
        if 'min' in cfg['rank_latest'] and rank_latest < cfg['rank_latest']['min']:
            return False
        if 'max' in cfg['rank_latest'] and rank_latest > cfg['rank_latest']['max']:
            return False
    
    if 'rank_change_rate' in cfg:
        if 'min' in cfg['rank_change_rate'] and rank_change_rate < cfg['rank_change_rate']['min']:
            return False
        if 'max' in cfg['rank_change_rate'] and rank_change_rate > cfg['rank_change_rate']['max']:
            return False
    
    # 市场特征检查
    monthly_search = row.get('20260113月搜', 0)
    if 'monthly_search' in cfg:
        if 'min' in cfg['monthly_search'] and monthly_search < cfg['monthly_search']['min']:
            return False
        if 'max' in cfg['monthly_search'] and monthly_search > cfg['monthly_search']['max']:
            return False
    
    monthly_purchase = row.get('月购', 0)
    if 'monthly_purchase' in cfg:
        if 'min' in cfg['monthly_purchase'] and monthly_purchase < cfg['monthly_purchase']['min']:
            return False
        if 'max' in cfg['monthly_purchase'] and monthly_purchase > cfg['monthly_purchase']['max']:
            return False
    
    product_count = row.get('广告竞品数', 0)
    if 'product_count' in cfg:
        if 'max' in cfg['product_count'] and product_count > cfg['product_count']['max']:
            return False
        if 'min' in cfg['product_count'] and product_count < cfg['product_count']['min']:
            return False
    
    # 竞争特征检查
    conversion_conc = row.get('转化集中度', 0)
    if 'conversion_concentration' in cfg:
        if 'min' in cfg['conversion_concentration'] and conversion_conc < cfg['conversion_concentration']['min']:
            return False
        if 'max' in cfg['conversion_concentration'] and conversion_conc > cfg['conversion_concentration']['max']:
            return False
    
    click_conc = row.get('点击集中度', 0)
    if 'click_concentration' in cfg:
        if 'min' in cfg['click_concentration'] and click_conc < cfg['click_concentration']['min']:
            return False
        if 'max' in cfg['click_concentration'] and click_conc > cfg['click_concentration']['max']:
            return False
    
    pcp_price = row.get('PCP竞价（美元）', 0)
    if 'pcp_price' in cfg:
        if 'min' in cfg['pcp_price'] and pcp_price < cfg['pcp_price']['min']:
            return False
        if 'max' in cfg['pcp_price'] and pcp_price > cfg['pcp_price']['max']:
            return False
    
    # SPR检查（维持首页排名的8天预估单量）
    spr = row.get('SPR', 0)
    if 'spr' in cfg:
        if 'max' in cfg['spr'] and spr > cfg['spr']['max']:
            return False
    
    # 标题密度检查（关键词在标题中的密度）
    title_density = row.get('标题密度', 0)
    if 'title_density' in cfg:
        if 'max' in cfg['title_density'] and title_density > cfg['title_density']['max']:
            return False
    
    # 产品特征检查
    price = row.get('价格（美元）', 0)
    if 'price' in cfg:
        if 'min' in cfg['price'] and price < cfg['price']['min']:
            return False
        if 'max' in cfg['price'] and price > cfg['price']['max']:
            return False
    
    review_count = row.get('评分数', 0)
    if 'review_count' in cfg:
        if 'min' in cfg['review_count'] and review_count < cfg['review_count']['min']:
            return False
        if 'max' in cfg['review_count'] and review_count > cfg['review_count']['max']:
            return False
    
    rating = row.get('评分值', 0)
    if 'rating' in cfg:
        if 'min' in cfg['rating'] and rating < cfg['rating']['min']:
            return False
        if 'max' in cfg['rating'] and rating > cfg['rating']['max']:
            return False
    
    need_supply_ratio = row.get('需供比', 0)
    if 'need_supply_ratio' in cfg:
        if 'min' in cfg['need_supply_ratio'] and need_supply_ratio < cfg['need_supply_ratio']['min']:
            return False
        if 'max' in cfg['need_supply_ratio'] and need_supply_ratio > cfg['need_supply_ratio']['max']:
            return False
    
    # 购买率检查
    purchase_rate = row.get('购买率', 0)
    if 'purchase_rate' in cfg:
        if 'min' in cfg['purchase_rate'] and purchase_rate < cfg['purchase_rate']['min']:
            return False
    
    # 毛利率检查（基于新的成本模型）
    price = row.get('价格（美元）', 0)
    estimated_margin = calculate_profit_margin(price)
    if estimated_margin < 10:  # 毛利率<10%的过滤掉
        return False
    
    return True


# ============================================================
# 主程序
# ============================================================

# 测试模式：只处理前N个关键词
TEST_MODE = False  # True=测试模式，False=全量运行
TEST_COUNT = 99    # 测试模式处理的关键词数量

# LLM分析开关
LLM_ANALYSIS_ENABLED = True  # True=启用LLM分析（火山方舟 ark-code-latest），False=使用关键词匹配

# LLM分析缓存（避免重复调用）
llm_cache = {}

# 数据源配置
# 测试版本：使用旧的选品结果（快速测试脚本）
# 正式版本：使用原始关键词数据
DATA_SOURCE = {
    "test": {
        "file": os.path.join(BASE_DIR, "output", "dummy_test.xlsx"),
        "description": "测试模式：使用旧的选品结果"
    },
    "production": {
        "file": None,  # None表示自动从RAW_DATA_DIR查找最新的关键词文件
        "description": "正式模式：使用原始关键词数据"
    }
}

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 80)
    print("选品模块 - 完整版（16个核心指标 + 5W1H场景还原 + LLM分析）")
    print("=" * 80)
    print(f"\n模式: {'测试' if TEST_MODE else '正式'}")
    print(f"LLM分析: {'开启' if LLM_ANALYSIS_ENABLED else '关闭'}")
    print("\n核心标准:")
    print("   1. 价格$10-50（冲动购买甜蜜点）")
    print("   2. 月销300-900（10×10×1法则: 每天10单×$10利润）")
    print("   3. 竞品评论<10,000（避免饱和市场）")
    print("   4. 毛利率>30%红线")
    print("   5. 无大品牌垄断")
    print("   6. 非季节性产品（允许，但需AI分析窗口期）")
    print("   7. 市场深度: 前3产品<60%市场份额")
    print("=" * 80)

    # 1. 读取数据
    if TEST_MODE:
        # 测试模式：使用旧选品结果（已有计算好的列）
        data_file = DATA_SOURCE["test"]["file"]
        print(f"\n读取(测试模式): {os.path.basename(data_file)}")
        df = pd.read_excel(data_file)
        print(f"数据量: {len(df)} 条")
    else:
        # 正式模式：从原始数据目录查找最新关键词文件
        raw_files = [f for f in os.listdir(RAW_DATA_DIR) if '关键词' in f and f.endswith('.xlsx')]
        if not raw_files:
            print("未找到关键词数据文件")
            return
        # 优先选择 search_全部关键词 文件（含完整维度的那个）
        full_files = [f for f in raw_files if f.startswith('search')]
        if full_files:
            chosen = full_files[0]
        else:
            chosen = raw_files[0]
        data_file = os.path.join(RAW_DATA_DIR, chosen)
        print(f"\n读取(正式模式): {chosen}")
        df = pd.read_excel(data_file)

        # 数值转换
        numeric_cols = ['20260113月搜', '月购', '月展示', '月点击', '需供比', '点击集中度', '转化集中度',
                        'SPR', '标题密度', '广告竞品数', '评分数', '评分值', '价格（美元）', 
                        'PCP竞价（美元）', '购买率']
        
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 排名列
        rank_cols = ['20260106排名', '20260113排名', '20260120排名', '20260127排名']
        for col in rank_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 计算排名变化
        df['rank_earliest'] = df['20260106排名']
        df['rank_latest'] = df['20260127排名']
        df['rank_change'] = df['rank_earliest'] - df['rank_latest']
        df['rank_change_rate'] = df['rank_change'] / (df['rank_earliest'] + 1)
        
        # 计算毛利率
        df['预估毛利率'] = df['价格（美元）'].apply(calculate_profit_margin)
        
        # 清理无效数据
        df = df[(df['rank_earliest'] > 0) & (df['rank_latest'] > 0)].copy()
        df = df.dropna(subset=['20260113月搜', '广告竞品数', '价格（美元）'])
    
    print(f"有效数据: {len(df):,} 条")

    # 2. 按策略筛选或直接使用（测试模式跳过）
    if TEST_MODE:
        # 测试模式：直接使用数据，跳过筛选
        df_final = df.copy()
        # 取前N个关键词
        unique_kws = df_final['关键词(绿色建议进入，黄色找切入点，粉色观察)'].unique()[:TEST_COUNT]
        df_final = df_final[df_final['关键词(绿色建议进入，黄色找切入点，粉色观察)'].isin(unique_kws)]
        print(f"\n[TEST] 测试模式：取前{TEST_COUNT}个关键词 ({len(df_final)}条)")
    else:
        # 正式模式：执行4种策略筛选
        print("\n执行4种策略筛选...")
        all_selected = []
        seen_keywords = set()

        strategy_keys = [
            "strategy1_blue_ocean",
            "strategy2_red_ocean",
            "strategy3_differentiation",
            "strategy4_follow",
        ]

        for key in strategy_keys:
            cfg = STRATEGY_CONFIG[key]
            name = cfg["name"]
            quota = cfg["quota"]

            # 筛选
            filtered = df[df.apply(lambda row: check_strategy_match(row, key), axis=1)]
            # 去重
            filtered = filtered[~filtered['关键词(绿色建议进入，黄色找切入点，粉色观察)'].isin(seen_keywords)]
            # 按排名变化排序（优先选排名上升的）
            filtered = filtered.sort_values('rank_change', ascending=False)
            # 取top
            taken = filtered.head(quota)

            if len(taken) > 0:
                taken = taken.copy()
                taken['分配策略'] = name
                all_selected.append(taken)
                seen_keywords.update(taken['关键词(绿色建议进入，黄色找切入点，粉色观察)'].tolist())

            print(f"  {name}: {len(taken)}/{quota} 个")

        # 合并
        if not all_selected:
            print("\n筛选结果为空")
            return

        df_final = pd.concat(all_selected, ignore_index=True)
        df_final = df_final.reset_index(drop=True)

        # 确保不超过total_quota
        max_quota = STRATEGY_CONFIG.get("total_quota", 100)
        if len(df_final) > max_quota:
            df_final = df_final.head(max_quota)

    # 3. 分析每个产品
    print(f"\n{'=' * 80}")
    print("16个核心指标检测 + 5W1H场景还原...")
    print(f"{'=' * 80}")
    
    df_final['季节性'] = ''
    df_final['热点事件'] = ''
    df_final['品牌竞争'] = ''
    df_final['分析理由'] = ''
    df_final['预估毛利率'] = 0.0
    df_final['预估利润'] = 0.0
    df_final['指标通过率'] = ''
    df_final['用户场景'] = ''
    df_final['类别优先级'] = ''
    df_final['类别备注'] = ''

    for idx, row in df_final.iterrows():
        keyword = str(row['关键词(绿色建议进入，黄色找切入点，粉色观察)'])
        
        # 大模型分析（使用缓存）
        if keyword in llm_cache:
            analysis = llm_cache[keyword]
        else:
            analysis = call_llm_for_analysis(keyword, row)
            llm_cache[keyword] = analysis
        
        # 季节性
        if analysis.get('is_seasonal'):
            df_final.at[idx, '季节性'] = '⚠️季节'
        else:
            df_final.at[idx, '季节性'] = '✅'
        
        # 热点事件
        if analysis.get('is_trending'):
            df_final.at[idx, '热点事件'] = '🔥热点'
        else:
            df_final.at[idx, '热点事件'] = '✅'
        
        # 品牌竞争
        if analysis.get('has_brand_competition'):
            df_final.at[idx, '品牌竞争'] = '⚠️品牌'
        else:
            df_final.at[idx, '品牌竞争'] = '✅'
        
        # 综合分析理由（详细版）
        reasons_parts = []
        if analysis.get('is_seasonal'):
            reasons_parts.append(f"季节:{analysis.get('seasonal_window', '季节性产品')}")
        if analysis.get('is_trending'):
            reasons_parts.append(f"热点:{analysis.get('trending_future', '热点驱动')}")
        if analysis.get('has_brand_competition'):
            reasons_parts.append(f"品牌:{analysis.get('brand_window', '存在竞争')}")
        
        if reasons_parts:
            df_final.at[idx, '分析理由'] = ' | '.join(reasons_parts)
        else:
            df_final.at[idx, '分析理由'] = '✅常规产品' if not analysis.get('overall_reason') else analysis.get('overall_reason', '✅常规产品')
        
        # 毛利率和利润（基于新的成本模型）
        price = row['价格（美元）']
        margin = calculate_profit_margin(price)
        profit = calculate_profit_per_unit(price, margin)
        df_final.at[idx, '预估毛利率'] = margin
        df_final.at[idx, '预估利润'] = profit
        
        # 16指标检测
        indicators = check_16_indicators(row)
        df_final.at[idx, '指标通过率'] = f"{indicators['pass_rate']:.0f}%"
        
        # 5W1H场景
        monthly_search = row.get('20260113月搜', 0)
        review_count = row.get('评分数', 0)
        rating = row.get('评分值', 0)
        scene = analyze_5w1h_scene(keyword, monthly_search, price, review_count, rating)
        df_final.at[idx, '用户场景'] = scene['who'] + "/" + scene['where']
        
        # 类别优先级
        priority, note = get_category_priority(keyword)
        df_final.at[idx, '类别优先级'] = priority
        df_final.at[idx, '类别备注'] = note

    # 过滤掉食品类
    food_count = (df_final['类别优先级'] == '❌过滤').sum()
    if food_count > 0:
        print(f"\n🚫 过滤食品类产品: {food_count} 个")
        df_final = df_final[df_final['类别优先级'] != '❌过滤'].copy()

    # 4. 统计
    print(f"\n{'=' * 80}")
    print(f"最终精选: {len(df_final)} 个")
    print(f"{'=' * 80}")

    print("\n【策略分布】")
    for s, c in df_final['分配策略'].value_counts().items():
        pct = c / len(df_final) * 100
        print(f"  {s}: {c} ({pct:.0f}%)")

    print("\n【季节性分布】")
    seasonal = (df_final['季节性'] == '⚠️季节').sum()
    print(f"  季节性产品: {seasonal} 个 ({seasonal/len(df_final)*100:.0f}%)")

    print("\n【毛利率分布】")
    margin_above_25 = (df_final['预估毛利率'] >= 25).sum()
    margin_above_20 = (df_final['预估毛利率'] >= 20).sum()
    print(f"  毛利率≥25%: {margin_above_25} 个 ({margin_above_25/len(df_final)*100:.0f}%)")
    print(f"  毛利率≥20%: {margin_above_20} 个 ({margin_above_20/len(df_final)*100:.0f}%)")
    avg_margin = df_final['预估毛利率'].mean()
    avg_profit = df_final['预估利润'].mean()
    print(f"  平均毛利率: {avg_margin:.1f}%")
    print(f"  平均利润: ${avg_profit:.2f}/件")

    print("\n【16指标通过率】")
    avg_pass_rate = df_final['指标通过率'].str.replace('%', '').astype(float).mean()
    print(f"  平均通过率: {avg_pass_rate:.0f}%")

    print("\n【类别优先级分布】")
    for p, c in df_final['类别优先级'].value_counts().items():
        pct = c / len(df_final) * 100
        print(f"  {p}: {c} ({pct:.0f}%)")

    # 5. 保存（Times New Roman字体 + 首行冻结）
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(OUTPUT_DIR, f"精选选品_{len(df_final)}条_{timestamp}.xlsx")

    save_cols = [
        '关键词(绿色建议进入，黄色找切入点，粉色观察)', 
        '20260113月搜', '月购', '广告竞品数', '价格（美元）',
        '评分数', '评分值', '转化集中度', '点击集中度', '需供比', 'PCP竞价（美元）',
        'SPR', '标题密度',  # 新增
        '分配策略', 'rank_earliest', 'rank_latest', 'rank_change_rate',
        '季节性', '热点事件', '品牌竞争', '分析理由',
        '预估毛利率', '预估利润', '指标通过率', '用户场景',
        '类别优先级', '类别备注',
    ]
    save_cols = [c for c in save_cols if c in df_final.columns]
    
    # 保存为Excel并设置格式
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_final[save_cols].to_excel(writer, index=False, sheet_name='精选选品')
        worksheet = writer.sheets['精选选品']
        
        # 首行冻结
        worksheet.freeze_panes = 'A2'
        
        # 设置Times New Roman字体
        header_font = Font(name='Times New Roman', size=11, bold=True)
        data_font = Font(name='Times New Roman', size=10)
        
        # 格式化表头
        for cell in worksheet[1]:
            cell.font = header_font
        
        # 格式化数据
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"\n✅ 已保存: {output_file}")

    # 6. 显示Top 30
    print(f"\n{'=' * 80}")
    print("Top 30 精选结果（按排名变化排序）")
    print(f"{'=' * 80}")

    # 排序列（兼容不同数据源）
    sort_col = 'rank_change' if 'rank_change' in df_final.columns else 'rank_change_rate'
    df_show = df_final.sort_values(sort_col, ascending=False).head(30)
    for i, (_, row) in enumerate(df_show.iterrows()):
        kw = str(row['关键词(绿色建议进入，黄色找切入点，粉色观察)'])[:35]
        margin = row['预估毛利率']
        profit = row['预估利润']
        pass_rate = row['指标通过率']
        priority = row['类别优先级']
        note = row['类别备注']
        spr = row.get('SPR', 0)
        
        print(f"\n{i+1:2d}. {kw}")
        print(f"    {row['分配策略']} | {priority} {note}")
        print(f"    排名:{int(row['rank_earliest']):,}→{int(row['rank_latest']):,} ({row['rank_change_rate']:+.1%})")
        print(f"    月搜:{int(row['20260113月搜']):,} | 竞品:{int(row['广告竞品数']):,} | SPR:{int(spr)} | ${row['价格（美元）']:.2f} | 毛利:{margin:.1f}%(${profit:.2f})")
        print(f"    评分:{row['评分值']}★ | 评论:{int(row['评分数']):,} | 通过率:{pass_rate}")


if __name__ == "__main__":
    main()
